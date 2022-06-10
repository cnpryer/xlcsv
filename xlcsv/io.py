import csv
import os
from io import BytesIO, StringIO
from pathlib import Path
from typing import BinaryIO, Optional, Union

from openpyxl import load_workbook  # type: ignore

from xlcsv import utils
from xlcsv.error import OpenpyxlException

# default .xlsx sheet name
DEFAULT_EXCEL_SHEET_NAME = "Sheet1"

# default openpyxl config
DEFAULT_OPENPYXL_OPTIONS = {"read_only": True, "data_only": True}

# supported file extensions
EXCEL_FILE_EXTENSIONS = {".xlsx"}


def excel_to_csv_buffer(
    file_like: Union[str, BytesIO, Path, BinaryIO, bytes],
    sheet_name: Optional[str] = None,
    sheet_index: Optional[int] = None,
    openpyxl_options: Optional[dict] = dict(DEFAULT_OPENPYXL_OPTIONS),
) -> StringIO:
    """Build a reset StringIO buffer of CSV data from Excel.

    Args:
        file_like (Union[str, BytesIO, Path, BinaryIO, bytes]):
            Path to a file or a file-like object. Objects with a
            ``read()`` method, such as a file handler
            (e.g. via builtin ``open`` function) or ``BytesIO``.
        sheet_name (Optional[str], optional): Name of sheet to read.
            Defaults to None.
        sheet_index (Optional[int], optional): Position of sheet in book.
            Defaults to None.
        openpyxl_options (Optional[dict], optional): Extra openpyxl options
            for parsing Excel files.

    Returns:
        StringIO: Reset String IO buffer.
    """

    if isinstance(file_like, (str, Path)):
        file_like = utils.format_path(file_like)
        file_ext = os.path.splitext(file_like)[1]

        if file_ext not in EXCEL_FILE_EXTENSIONS:
            raise OpenpyxlException("File extension not allowed.")

    # create book
    book = load_workbook(file_like, **openpyxl_options)
    sheet = book[sheet_name or sheet_index or DEFAULT_EXCEL_SHEET_NAME]

    # write book to csv string io
    buffer = StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)
    for row in sheet.iter_rows():
        data = []
        for cell in row:
            data.append(cell.value)
        writer.writerow(data)

    # reset buffer
    buffer.seek(0)

    return buffer
