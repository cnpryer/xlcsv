from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook  # type: ignore

from xlcsv import utils
from xlcsv.error import FileExtensionException

# Default .xlsx
DEFAULT_EXCEL_SHEET_NAME = "Sheet1"
DEFAULT_SHEET_INDEX = 0

# Default openpyxl config
DEFAULT_OPENPYXL_OPTIONS = {"read_only": True, "data_only": True}

# Supported file extensions
EXCEL_FILE_EXTENSIONS = {".xlsx"}


def to_csv_buffer(
    file_like: str | BytesIO | Path | BinaryIO | bytes,
    sheet_name: str | None = None,
    sheet_index: int | None = None,
    openpyxl_options: dict | None = dict(DEFAULT_OPENPYXL_OPTIONS),
) -> StringIO:
    """Build a reset StringIO buffer of CSV data from Excel.

    Args:
        file_like (str | BytesIO | Path | BinaryIO | bytes):
            Path to a file or a file-like object. Objects with a
            ``read()`` method, such as a file handler
            (e.g. via builtin ``open`` function) or ``BytesIO``.
        sheet_name (str | None, optional): Name of sheet to read.
            Defaults to None.
        sheet_index (int | None, optional): Position of sheet in book.
            Defaults to None.
        openpyxl_options (dict | None, optional): Extra openpyxl options
            for parsing Excel files.
    Returns:
        StringIO: Reset String IO buffer.
    """
    if isinstance(file_like, str):
        file_like = Path(file_like)

    if isinstance(file_like, Path):  # type: ignore
        file_like = utils.format_path(file_like)

        if file_like.suffix not in EXCEL_FILE_EXTENSIONS:
            raise FileExtensionException("File extension not allowed.")

        file_like = file_like.as_posix()

    # Create book
    book = load_workbook(file_like, **openpyxl_options)
    sheet = book[sheet_name or sheet_index or DEFAULT_EXCEL_SHEET_NAME]

    # Write book to csv string io
    buffer = StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)
    for row in sheet.iter_rows():
        data = [cell.value for cell in row]
        writer.writerow(data)

    # Reset buffer
    buffer.seek(0)

    return buffer
